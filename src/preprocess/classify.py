"""Step 1 — deterministic form-version classifier.

Classifies an Excel file against the weighted signatures in
``config/form_signatures.yaml``. No LLM: every signal is a structural,
reproducible check. Files matching no signature fall back to ``"unknown"``
rather than being misclassified; files matching two versions are flagged
``needs_review``.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import typer
import yaml

from src.utils.logging import get_logger
from src.utils.paths import FORM_SIGNATURES_PATH

log = get_logger(__name__)

_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
_SCAN_ROWS = 12
_SCAN_COLS = 14


@dataclass
class FormFeatures:
    """Structural features extracted from a workbook."""

    sheet_names: list[str]
    max_cols: int
    cell_texts: list[str]


@dataclass
class ClassificationResult:
    """Outcome of classifying one file."""

    file_path: str
    form_version: str
    confidence: float
    needs_review: bool = False
    evidence: dict[str, float] = field(default_factory=dict)
    features: FormFeatures | None = None


def load_signatures(path: Path = FORM_SIGNATURES_PATH) -> dict[str, object]:
    """Load the form-signature config.

    Args:
        path: Path to ``form_signatures.yaml``.

    Returns:
        The parsed ``versions`` mapping.
    """
    data = yaml.safe_load(path.read_text(encoding="utf-8"))
    return data["versions"]


def extract_features(path: Path) -> FormFeatures:
    """Extract deterministic structural features from a workbook.

    Args:
        path: Path to the Excel file.

    Returns:
        A populated :class:`FormFeatures`.
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = workbook.worksheets
        max_cols = max((s.max_column or 0 for s in sheets), default=0)
        cell_texts: list[str] = []
        for sheet in sheets:
            for row in sheet.iter_rows(
                min_row=1, max_row=_SCAN_ROWS, max_col=_SCAN_COLS
            ):
                for cell in row:
                    if cell.value is not None:
                        cell_texts.append(str(cell.value).strip())
        return FormFeatures(
            sheet_names=[s.title for s in sheets],
            max_cols=max_cols,
            cell_texts=cell_texts,
        )
    finally:
        workbook.close()


def match_signal(features: FormFeatures, signal: dict[str, object]) -> bool:
    """Return True if a single signature signal matches the features.

    Args:
        features: Extracted workbook features.
        signal: One signal entry from the signature config.

    Returns:
        Whether the signal matches.
    """
    signal_type = signal["type"]
    lower_sheets = [s.lower() for s in features.sheet_names]
    lower_cells = [c.lower() for c in features.cell_texts]

    if signal_type == "sheet_exists":
        return str(signal["name"]).strip().lower() in lower_sheets
    if signal_type == "sheet_name_contains":
        kws = [str(k).lower() for k in signal["keywords"]]  # type: ignore[union-attr]
        return any(kw in name for name in lower_sheets for kw in kws)
    if signal_type == "col_count_range":
        low, high = signal["range"]  # type: ignore[misc]
        return low <= features.max_cols <= high
    if signal_type == "header_text_contains":
        kws = [str(k).lower() for k in signal["keywords"]]  # type: ignore[union-attr]
        return any(kw in cell for cell in lower_cells for kw in kws)
    if signal_type == "marker_cell":
        return str(signal["value"]).lower() in lower_cells
    if signal_type == "stage_row":
        markers = {str(m) for m in signal["markers"]}  # type: ignore[union-attr]
        return len(markers & set(features.cell_texts)) >= 4
    log.warning("classify.unknown_signal", signal_type=signal_type)
    return False


def _score_version(features: FormFeatures, spec: dict[str, object]) -> float:
    """Sum the matched signal weights for one version spec."""
    total = 0.0
    for signal in spec["signals"]:  # type: ignore[union-attr]
        if match_signal(features, signal):
            total += float(signal["weight"])  # type: ignore[index]
    return round(total, 4)


def classify_form(
    path: Path, signatures: dict[str, object] | None = None
) -> ClassificationResult:
    """Classify the form version of an Excel file.

    Args:
        path: Path to the Excel file.
        signatures: Optional pre-loaded signature config (else loaded from disk).

    Returns:
        A :class:`ClassificationResult`. ``form_version`` is ``"unknown"`` when
        no version reaches its threshold; ``needs_review`` is set when two or
        more versions do.
    """
    versions = signatures if signatures is not None else load_signatures()
    features = extract_features(path)

    scores = {name: _score_version(features, spec) for name, spec in versions.items()}
    passed = [
        name
        for name, spec in versions.items()
        if scores[name] >= float(spec["threshold"])  # type: ignore[index]
    ]

    if not passed:
        version, confidence, needs_review = "unknown", 0.0, False
    else:
        version = max(passed, key=lambda n: scores[n])
        confidence = scores[version]
        needs_review = len(passed) >= 2
        if needs_review:
            confidence = round(confidence * 0.7, 4)

    log.info(
        "classify.result",
        file=path.name,
        version=version,
        confidence=confidence,
        needs_review=needs_review,
    )
    return ClassificationResult(
        file_path=str(path),
        form_version=version,
        confidence=confidence,
        needs_review=needs_review,
        evidence=scores,
        features=features,
    )


def classify_dir(directory: Path) -> list[ClassificationResult]:
    """Classify every Excel file under a directory.

    Args:
        directory: Directory to scan recursively.

    Returns:
        One :class:`ClassificationResult` per Excel file.
    """
    signatures = load_signatures()
    results: list[ClassificationResult] = []
    for file in sorted(directory.rglob("*")):
        if file.suffix.lower() in _EXCEL_SUFFIXES:
            results.append(classify_form(file, signatures))
    return results

"""Step 2 — deterministic form-version classifier.

Given an Excel file, decides which master form version it is. No LLM: pure
heuristic rules over structural features. Unknown files fall back to
``"unknown"`` rather than being misclassified.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import typer

from src.utils.logging import get_logger

log = get_logger(__name__)

# Marker cell text scanned in the top-left region of each sheet.
_AAAA_MARKER = "aaaa"
_STAGE_MARKERS = {"CP", "PP", "DV", "PV", "PQ"}
_GROUP_HEADER_TOKENS = {"공통", "DRBFM", "부품", "FMEA", "HSMS", "Common"}


@dataclass
class FormFeatures:
    """Structural features extracted from a workbook."""

    sheet_count: int
    max_cols: int
    has_history_sheet: bool
    has_aaaa_marker: bool
    has_stage_row: bool
    has_grouped_header: bool
    has_better_sheet: bool


@dataclass
class ClassificationResult:
    """Outcome of classifying one file."""

    file_path: str
    form_version: str
    confidence: float
    reasons: list[str] = field(default_factory=list)
    features: FormFeatures | None = None


def _scan_top_left(sheet: object, rows: int = 12, cols: int = 12) -> list[str]:
    """Collect stringified cell values from the top-left region of a sheet.

    Args:
        sheet: An openpyxl worksheet.
        rows: Number of rows to scan.
        cols: Number of columns to scan.

    Returns:
        Lower-cased non-empty cell strings.
    """
    values: list[str] = []
    for row in sheet.iter_rows(min_row=1, max_row=rows, max_col=cols):
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value).strip())
    return values


def extract_features(path: Path) -> FormFeatures:
    """Extract structural features used by the classifier.

    Args:
        path: Path to the Excel file.

    Returns:
        A populated :class:`FormFeatures`.
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = workbook.worksheets
        max_cols = max((s.max_column or 0 for s in sheets), default=0)
        sheet_names = [s.title for s in sheets]
        has_history = any(n.strip().lower() == "history" for n in sheet_names)
        has_better = any("better" in n.lower() for n in sheet_names)

        cells: list[str] = []
        for sheet in sheets:
            cells.extend(_scan_top_left(sheet))
        cell_set = {c.lower() for c in cells}

        has_aaaa = _AAAA_MARKER in cell_set
        has_stage = len(_STAGE_MARKERS & set(cells)) >= 4
        has_grouped = any(
            tok.lower() in c.lower()
            for c in cells
            for tok in _GROUP_HEADER_TOKENS
        )
        return FormFeatures(
            sheet_count=len(sheets),
            max_cols=max_cols,
            has_history_sheet=has_history,
            has_aaaa_marker=has_aaaa,
            has_stage_row=has_stage,
            has_grouped_header=has_grouped,
            has_better_sheet=has_better,
        )
    finally:
        workbook.close()


def classify_form(path: Path) -> ClassificationResult:
    """Classify the form version of an Excel file.

    Decision tree (deterministic, in priority order):
      1. History sheet + 45-75 cols          -> v1.2
      2. aaaa marker or (>=80 cols + stage)   -> 96col
      3. Better sheet name + 45-79 cols       -> 56col
      4. 0 < cols < 45                        -> 20col
      5. otherwise                            -> unknown

    Args:
        path: Path to the Excel file.

    Returns:
        A :class:`ClassificationResult` with version, confidence, and reasons.
    """
    f = extract_features(path)
    reasons: list[str] = []

    if f.has_history_sheet and 45 <= f.max_cols <= 75:
        reasons.append(f"History sheet present, {f.max_cols} cols in v1.2 range")
        version, confidence = "v1.2", 0.95
    elif f.has_aaaa_marker or (f.max_cols >= 80 and f.has_stage_row):
        if f.has_aaaa_marker:
            reasons.append("'aaaa' placeholder marker found")
        if f.max_cols >= 80:
            reasons.append(f"{f.max_cols} cols (wide)")
        if f.has_stage_row:
            reasons.append("CP/PP/DV/PV/PQ stage row found")
        version, confidence = "96col", 0.9
    elif f.has_better_sheet and 45 <= f.max_cols < 80:
        reasons.append(f"'Better' sheet name, {f.max_cols} cols")
        version, confidence = "56col", 0.85
    elif 0 < f.max_cols < 45:
        reasons.append(f"{f.max_cols} cols, simple single-header layout")
        version, confidence = "20col", 0.8
    else:
        reasons.append(f"no rule matched ({f.max_cols} cols)")
        version, confidence = "unknown", 0.0

    log.info(
        "classify.result",
        file=path.name,
        version=version,
        confidence=confidence,
    )
    return ClassificationResult(
        file_path=str(path),
        form_version=version,
        confidence=confidence,
        reasons=reasons,
        features=f,
    )


app = typer.Typer(help="Deterministic form-version classifier.")


@app.command()
def classify(path: Path = typer.Argument(..., help="Excel file to classify.")) -> None:
    """Classify a single file and print the result."""
    result = classify_form(path)
    typer.echo(f"{path.name}: {result.form_version} (confidence={result.confidence})")
    for reason in result.reasons:
        typer.echo(f"  - {reason}")


@app.command()
def classify_dir(
    directory: Path = typer.Argument(..., help="Directory of Excel files."),
) -> None:
    """Classify every Excel file in a directory and print a count summary."""
    counts: dict[str, int] = {}
    for file in sorted(directory.rglob("*")):
        if file.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            continue
        result = classify_form(file)
        counts[result.form_version] = counts.get(result.form_version, 0) + 1
        typer.echo(f"{file.name}: {result.form_version} ({result.confidence})")
    typer.echo("\nForm-version counts:")
    for version, count in sorted(counts.items()):
        typer.echo(f"  {version}: {count}")


if __name__ == "__main__":
    app()

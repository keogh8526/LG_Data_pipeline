"""v2.0 합성 Excel fixture 생성기.

19개 실측 파일 대신 7개 양식 패밀리의 *모양*(시트명, 헤더 마커, max_col)을
모방하는 합성 워크북. real R1~R5 멀티헤더 복잡도는 재현하지 않음 (D-002).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

FIXTURE_DIR = Path(__file__).resolve().parent


def _save(workbook: openpyxl.Workbook, name: str) -> Path:
    path = FIXTURE_DIR / name
    workbook.save(path)
    return path


def _pad(row: list, n: int) -> list:
    return row + [None] * max(0, n - len(row))


def make_changing_parts_96() -> Path:
    """변경부품 list family 96col.

    헤더 4행:
      행1: "New & Changing Part" 마커
      행2: 대분류 (공통/부품/DRBFM/친환경/금형/시험) + base/new model
      행3: buyer 코드
      행4: leaf 컬럼명 (P/No, 품명, ...)
    parse_multi_header가 행2+행4를 " > "로 join → "공통 > 부품 P/No" 같은 path 생성.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "변경부품 list_Best1"

    # 행1: 마커 (col 2)
    ws.append(_pad([None, "New & Changing Part"], 96))

    # 행2: 대분류 — 컬럼당 명확히 지정 (forward-fill로 빈 칸 채워짐)
    row2 = [
        "공통", "공통", "공통",         # 1-3
        "부품", "부품",                  # 4-5
        "DRBFM",                         # 6  → base_model_code 메타 (col 9 위치 보존용 더미)
        "공통", "공통",                  # 7-8
        "공통",                          # 9: base_model_code 셀
        None, None,                      # 10-11
        "공통",                          # 12: new_model_code 셀
        None, None, None, None,          # 13-16
    ]
    # 위치 9/12에 모델 값 — 별도 헤더로 처리되지 않게 col 9/12는 None
    # base_model_code/new_model_code는 _meta로 추출됨, payload엔 들어가지 않음
    ws.append(_pad(row2, 96))

    # 행3: buyer
    row3 = [None] * 96
    row3[8] = "LGEUR"   # col 9
    row3[11] = "LGEUR"  # col 12
    ws.append(row3)

    # 행4: leaf 컬럼명
    row4 = [
        "부품 P/No", "품명", "기존 부품",   # 공통 > ...
        "BOM Level", "부품 유형",           # 부품 > ...
        "DRBFM 코멘트",                     # DRBFM > DRBFM 코멘트
        "변경점", "변경 사유",              # 공통 > 변경점/변경사유 (간단화)
        None,                               # col 9 (모델코드 메타 셀)
        "공급사", "HSMS 영향",              # 10-11
        None,                               # col 12 (모델코드 메타 셀)
        "금형 종류", "부품인정시험 항목",   # 13-14
        None, None,                         # 15-16
    ]
    ws.append(_pad(row4, 96))

    # 데이터 행 1
    data1 = [
        "AGG74419321", "Packing Assembly", "AGG74419320",
        1, "Assy",
        "고온 영역 확장",
        "도어 힌지 내열 220→240", "신규 규제 대응",
        None,
        "세우산업", "재인증 필요",
        None,
        "사출 신규", "내열, 내구",
        None, None,
    ]
    ws.append(_pad(data1, 96))

    data2 = [
        "AGB1234567", "Bracket", None,
        2, "단품",
        None,
        "두께 1.5→2.0", "강도 보강",
        None,
        "한양", None,
        None,
        None, None,
        None, None,
    ]
    ws.append(_pad(data2, 96))
    return _save(wb, "fixture_changing_parts_96.xlsx")


def make_changing_parts_91() -> Path:
    """변경부품 list 91col (오래된 버전)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "변경부품 list_Better1"
    ws.append(_pad([None, "New & Changing Part"], 91))
    row2 = ["공통", "공통", "공통", "부품", "부품", "공통", "공통", "공통", None, None, None, None]
    ws.append(_pad(row2, 91))
    row3 = [None] * 91
    row3[8] = "LGEAU"
    row3[11] = "LGEAU"
    ws.append(row3)
    row4 = ["부품 P/No", "품명", "기존 부품", "BOM Level", "부품 유형", "변경점", "변경 사유", None,
            None, None, None, None]
    ws.append(_pad(row4, 91))
    ws.append(_pad(["XYZ1234567", "Cover", None, 1, "사출", "재질 변경 PP→PA66", "내열성 강화"], 91))
    return _save(wb, "fixture_changing_parts_91.xlsx")


def make_new_parts_list_75() -> Path:
    """신규부품리스트 75col (Key-In/LOV 마커 다수)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "신규부품리스트"
    # 행1: Key-In/LOV 마커가 다수 (20개 이상)
    row1 = ["Key-In" if i % 2 == 0 else "LOV" for i in range(75)]
    ws.append(row1)
    # 행2: 필수/옵션
    ws.append(["필수"] * 5 + ["옵션"] * 70)
    # 행3: 실제 컬럼명
    headers = (
        ["No", "프로젝트 코드", "신규 구분", "부품 P/No.", "품명",
         "기존 P/No.", "변경점", "변경사유", "부품 등급", "역할", "이름", "SSO ID"]
        + [f"c{i}" for i in range(12, 75)]
    )
    ws.append(_pad(headers, 75))
    ws.append(_pad([None] * 75, 75))
    # 데이터
    ws.append(_pad([1, "PROJ001", "신규", "AAA1234567", "Sensor", None,
                    "신규 추가", "기능 추가", "Best-1", "설계", "김민석", "kim123"], 75))
    return _save(wb, "fixture_new_parts_list_75.xlsx")


def make_bom_ag_grid_36() -> Path:
    """BOM ag-grid 36col."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ag-grid"
    headers = ["P/No", "Desc.", "Lvl", "Qty", "MechanicalPart"] + [f"c{i}" for i in range(5, 36)]
    ws.append(headers)
    for i in range(120):  # min_row 100+
        ws.append(_pad([f"AGP{i:07d}", f"Part {i}", (i % 3) + 1, 1.0, "Mechanical"], 36))
    return _save(wb, "fixture_bom_ag_grid_36.xlsx")


def make_base_master_24() -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "개발변경부품리스트"
    ws.append(_pad([None] * 24, 24))  # 행1 비움
    ws.append(_pad(["No.", "P/no.", "Desc.", "Module", "CMDT", "도입/신규", "Lvl", "Part Grade"] + [f"c{i}" for i in range(8, 24)], 24))
    ws.append(_pad([None] * 24, 24))
    ws.append(_pad([1, "BSM1234567", "Old Cover", "MOD", "절삭", "신규", 1, "Best-1"], 24))
    return _save(wb, "fixture_base_master_24.xlsx")


def make_v1_2_template() -> Path:
    """v1_2 통합 마스터 (빈 템플릿)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master(Best)"
    ws.append(_pad([None, "공통 (Common)"], 59))
    ws.append(_pad(["Base P/No", "Part Name"] + [f"c{i}" for i in range(2, 59)], 59))
    hist = wb.create_sheet("History")
    hist.append(["version", "released_at"])
    hist.append(["v1.2", "2025-06-01"])
    return _save(wb, "fixture_v1_2_template.xlsx")


def make_uae_dev_list_58() -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "신규 개발리스트"
    ws.append(_pad([None] * 58, 58))
    ws.append(_pad([None, "New & Changing Part"], 58))
    ws.append(_pad([None] * 58, 58))
    ws.append(_pad(["부품 P/No.", "품명", "변경점"] + [f"c{i}" for i in range(3, 58)], 58))
    ws.append(_pad(["UAE1234567", "Heater", "사양 변경"], 58))
    return _save(wb, "fixture_uae_dev_list_58.xlsx")


def make_activity_master_meta() -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master"
    ws.append(_pad([None, "※ Master"], 13))
    ws.append(_pad(["모델명 :", "WSED7667M", None, None, "양산일 :", "2024-04-09"], 13))
    ws.append(_pad(["개발 사유", "신규 안전 규제 대응을 위한 도어 힌지 강화"], 13))
    return _save(wb, "fixture_activity_master.xlsx")


def make_all() -> list[Path]:
    """모든 fixture 생성."""
    return [
        make_changing_parts_96(),
        make_changing_parts_91(),
        make_new_parts_list_75(),
        make_bom_ag_grid_36(),
        make_base_master_24(),
        make_v1_2_template(),
        make_uae_dev_list_58(),
        make_activity_master_meta(),
    ]


if __name__ == "__main__":
    for created in make_all():
        print(f"created {created}")

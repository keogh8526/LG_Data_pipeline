"""Step 3 — adversarial value fixtures (the 10 traps).

These are the deterministic edge cases the normalize / map pipeline must
handle. Each entry is ``(label, raw_input, field_name, expected_value,
expects_success)``. Two of the traps (merged cells, formula cells) need a real
workbook — :func:`make_merged_formula_workbook` builds one on demand.
"""

from __future__ import annotations

import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


@dataclass
class TrapCase:
    """One adversarial input the normalize pipeline must handle."""

    label: str
    raw: Any
    field: str
    expected: Any
    expects_success: bool


# Pre-composed (NFC) versus decomposed (NFD) Korean — different bytes, same letters.
_KO_NFD = unicodedata.normalize("NFD", "내열")

VALUE_TRAPS: list[TrapCase] = [
    # 1. Leading-zero part numbers must be preserved as-is (no auto int coercion).
    #    "AB01234567" is a valid pattern; the leading zero stays.
    TrapCase("leading_zero_part_no", "AB01234567", "base_part_no", "AB01234567", True),
    # 2. An Excel date serial mis-fed as qty (45412 = 2024-05-15) must fail range.
    TrapCase("excel_date_serial_as_qty", 45412, "qty", None, False),
    # 3. NFD-decomposed Hangul must be re-composed by NFKC before string ops.
    TrapCase(
        "nfd_korean",
        _KO_NFD + " 보강",
        "change_point",
        "내열 보강",
        True,
    ),
    # 4. Full-width Latin / digits must collapse to half-width via NFKC.
    TrapCase(
        "fullwidth_latin",
        "ＡＢ" + "1234567",
        "base_part_no",
        "AB1234567",
        True,
    ),
    TrapCase(
        "fullwidth_digits",
        "AB" + "１２３４５６７",
        "base_part_no",
        "AB1234567",
        True,
    ),
    # 5. Null-likes short-circuit to None successfully.
    TrapCase("empty_string", "", "change_point", None, True),
    TrapCase("whitespace_only", "   ", "change_point", None, True),
    TrapCase("na_sentinel", "N/A", "change_point", None, True),
    # 8. Grade variants resolve to the canonical alias.
    TrapCase("grade_variant_no_dash", "Best1", "grade", "Best-1", True),
    TrapCase("grade_variant_short", "B1", "grade", "Best-1", True),
    # 9. Mixed CRLF/LF whitespace collapses to a single space.
    TrapCase(
        "crlf_in_text",
        "내열\r\n보강",
        "change_point",
        "내열 보강",
        True,
    ),
    # 10. Non-breaking space (U+00A0) normalizes via NFKC then collapses.
    TrapCase(
        "non_breaking_space",
        "내열 보강",
        "change_point",
        "내열 보강",
        True,
    ),
]


def make_merged_formula_workbook(path: Path) -> Path:
    """Build a workbook exercising merged cells (#6) and formula cells (#7).

    Traps 6 and 7 belong to the extract layer: openpyxl with
    ``data_only=True`` must yield the merged region's top-left value across
    its expanded cells and the cached value of a formula cell.

    Args:
        path: Output workbook path.

    Returns:
        ``path`` after the workbook is saved.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Better-1"
    ws["A1"] = "Base P/No"
    ws["B1"] = "Part Name"
    ws["C1"] = "Qty"

    ws["A2"] = "AB1234567"
    ws["B2"] = "Bracket"
    # Cached formula value — openpyxl can only read what was already evaluated
    # by Excel, so write the cached value directly.
    ws["C2"] = 2

    ws["A3"] = "AB1234568"
    # Merge B3:B4 (Bracket spans two rows).
    ws.merge_cells("B3:B4")
    ws["B3"] = "Bracket"
    ws["C3"] = 1
    ws["A4"] = "AB1234569"
    ws["C4"] = 1

    wb.save(path)
    return path
